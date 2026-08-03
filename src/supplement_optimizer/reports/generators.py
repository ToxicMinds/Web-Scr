"""Report generators: Markdown, CSV, Excel, HTML and JSON.

Every report carries a generated-at timestamp. Outputs are written to a target
directory and the produced paths are returned so CI can upload/commit them.
"""

from __future__ import annotations

import json
from pathlib import Path

import polars as pl
from openpyxl import Workbook

from supplement_optimizer.optimizer import Solution
from supplement_optimizer.reports.builder import ReportData

_MAX_ROWS = 20
_TS_FMT = "%Y-%m-%d %H:%M:%S UTC"
_FLOAT_DP = 3


def _fmt(value: object) -> str:
    """Render a cell value, rounding floats for readable reports."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{round(value, _FLOAT_DP):g}"
    return str(value)


def _frame_to_md(frame: pl.DataFrame, *, limit: int = _MAX_ROWS) -> str:
    if frame.is_empty():
        return "_No data._"
    frame = frame.head(limit)
    header = "| " + " | ".join(frame.columns) + " |"
    divider = "| " + " | ".join("---" for _ in frame.columns) + " |"
    rows = ["| " + " | ".join(_fmt(v) for v in row) + " |" for row in frame.iter_rows()]
    return "\n".join([header, divider, *rows])


def _frame_to_html(frame: pl.DataFrame, *, limit: int = _MAX_ROWS) -> str:
    if frame.is_empty():
        return "<p><em>No data.</em></p>"
    frame = frame.head(limit)
    head = "".join(f"<th>{c}</th>" for c in frame.columns)
    body = "".join(
        "<tr>" + "".join(f"<td>{_fmt(v)}</td>" for v in row) + "</tr>" for row in frame.iter_rows()
    )
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def _solution_md(solution: Solution | None) -> str:
    if solution is None:
        return "**No feasible basket found.**"
    lines = [
        f"- **Total: {solution.total}**  ",
        f"- Strategy: {solution.strategy}  ",
        f"- Shipping confidence: {solution.shipping_confidence}  ",
        "",
    ]
    for sub in solution.sub_baskets:
        coupon = f", coupon `{sub.coupon_code}`" if sub.coupon_code else ""
        lines.append(
            f"### {sub.retailer_slug} — {sub.total} "
            f"(goods {sub.product_subtotal}, shipping {sub.shipping_cost}{coupon})"
        )
        for line in sub.lines:
            lines.append(
                f"- {line.quantity} × {line.offer.title} "  # noqa: RUF001
                f"@ {line.unit_price} = {line.line_total} "
                f"([link]({line.offer.url}))"
            )
        lines.append("")
    fulfilled = ", ".join(f"{k}: {v} g" for k, v in solution.fulfilled_g.items())
    lines.append(f"_Fulfilled: {fulfilled}_")
    return "\n".join(lines)


class ReportWriter:
    """Writes a :class:`ReportData` into every supported format."""

    def __init__(self, output_dir: Path) -> None:
        self._dir = output_dir
        self._dir.mkdir(parents=True, exist_ok=True)

    def write_all(self, data: ReportData) -> dict[str, Path]:
        """Write all reports and return a mapping of name -> path."""
        return {
            "markdown": self.write_markdown(data),
            "html": self.write_html(data),
            "json": self.write_json(data),
            "excel": self.write_excel(data),
            **self.write_csvs(data),
        }

    def write_markdown(self, data: ReportData) -> Path:
        ts = data.generated_at.strftime(_TS_FMT)
        dest = "/".join([data.request.destination_country])
        doc = f"""# Supplement Optimizer Report

_Generated: {ts}_
Destination: **{dest}** · Base currency: **{data.request.base_currency}**

## Best Basket

{_solution_md(data.best)}

## Retailer Rankings (whole basket, single retailer)

{_frame_to_md(data.retailer_rankings)}

## Cheapest Protein (by €/100 g protein)

{_frame_to_md(data.cheapest_protein.select(_protein_cols(data.cheapest_protein)))}

## Cheapest Creatine (by €/kg)

{_frame_to_md(data.cheapest_creatine.select(_creatine_cols(data.cheapest_creatine)))}
"""
        path = self._dir / "report.md"
        path.write_text(doc, encoding="utf-8")
        return path

    def write_html(self, data: ReportData) -> Path:
        ts = data.generated_at.strftime(_TS_FMT)
        body = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Supplement Optimizer Report</title>
<style>
 body{{font-family:system-ui,Arial,sans-serif;margin:2rem;color:#1a1a1a}}
 table{{border-collapse:collapse;margin:1rem 0;font-size:14px}}
 th,td{{border:1px solid #ddd;padding:6px 10px;text-align:left}}
 th{{background:#f4f4f4}} h1,h2{{margin-top:1.5rem}}
 .total{{font-size:1.4rem;font-weight:700;color:#0a7d33}}
</style></head><body>
<h1>Supplement Optimizer Report</h1>
<p><em>Generated: {ts}</em> · Destination: <strong>{data.request.destination_country}</strong></p>
<h2>Best Basket</h2>
<p class="total">{data.best.total if data.best else "No feasible basket"}</p>
{_best_basket_html(data.best)}
<h2>Retailer Rankings</h2>
{_frame_to_html(data.retailer_rankings)}
<h2>Cheapest Protein (€/100 g protein)</h2>
{_frame_to_html(data.cheapest_protein.select(_protein_cols(data.cheapest_protein)))}
<h2>Cheapest Creatine (€/kg)</h2>
{_frame_to_html(data.cheapest_creatine.select(_creatine_cols(data.cheapest_creatine)))}
</body></html>"""
        path = self._dir / "report.html"
        path.write_text(body, encoding="utf-8")
        return path

    def write_json(self, data: ReportData) -> Path:
        payload = {
            "generated_at": data.generated_at.isoformat(),
            "destination_country": data.request.destination_country,
            "base_currency": str(data.request.base_currency),
            "best_basket": json.loads(data.best.model_dump_json()) if data.best else None,
            "retailer_rankings": data.retailer_rankings.to_dicts(),
        }
        path = self._dir / "best_basket.json"
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        return path

    def write_csvs(self, data: ReportData) -> dict[str, Path]:
        outputs: dict[str, Path] = {}
        for name, frame in (
            ("rankings_csv", data.retailer_rankings),
            ("cheapest_protein_csv", data.cheapest_protein),
            ("cheapest_creatine_csv", data.cheapest_creatine),
            ("all_metrics_csv", data.all_metrics),
        ):
            path = self._dir / f"{name.removesuffix('_csv')}.csv"
            if frame.is_empty():
                path.write_text("", encoding="utf-8")
            else:
                frame.write_csv(path)
            outputs[name] = path
        return outputs

    def write_excel(self, data: ReportData) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet, frame in (
            ("Retailer Rankings", data.retailer_rankings),
            ("Cheapest Protein", data.cheapest_protein),
            ("Cheapest Creatine", data.cheapest_creatine),
            ("All Metrics", data.all_metrics),
        ):
            ws = workbook.create_sheet(sheet[:31])
            if frame.is_empty():
                ws.append(["No data"])
                continue
            ws.append(frame.columns)
            for row in frame.iter_rows():
                ws.append(list(row))
        path = self._dir / "report.xlsx"
        workbook.save(path)
        return path


def _protein_cols(frame: pl.DataFrame) -> list[str]:
    preferred = ["retailer", "title", "price", "price_per_kg", "price_per_100g_protein", "url"]
    return [c for c in preferred if c in frame.columns] or frame.columns


def _creatine_cols(frame: pl.DataFrame) -> list[str]:
    preferred = ["retailer", "title", "price", "pack_content_g", "price_per_kg", "url"]
    return [c for c in preferred if c in frame.columns] or frame.columns


def _best_basket_html(solution: Solution | None) -> str:
    if solution is None:
        return "<p>No feasible basket found.</p>"
    parts: list[str] = []
    for sub in solution.sub_baskets:
        items = "".join(
            f"<li>{line.quantity} × {line.offer.title} @ {line.unit_price} = {line.line_total}</li>"  # noqa: RUF001
            for line in sub.lines
        )
        parts.append(
            f"<h3>{sub.retailer_slug} — {sub.total}</h3>"
            f"<p>goods {sub.product_subtotal}, shipping {sub.shipping_cost}</p>"
            f"<ul>{items}</ul>"
        )
    return "".join(parts)
